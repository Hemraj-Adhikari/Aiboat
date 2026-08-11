import openpyxl, json, re, datetime
from collections import defaultdict

SRC = '/mnt/user-data/uploads/SEPTEMBER_2026_INTAKE_-_UNIVERSITY_DETAILS.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

def norm(s):
    if s is None: return ''
    s = str(s).upper()
    s = re.sub(r'ISC.*PATHWAY.*', '', s)
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    s = re.sub(r'\bUNIVERSITY\b', '', s)
    s = re.sub(r'\bOF\b', '', s)
    s = re.sub(r'\bTHE\b', '', s)
    s = re.sub(r'\bST\.?\b', 'SAINT', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

SKIP = {'HOME PAGE','UG_COURSES','PG_COURSES'}
sheet_norm_to_title = {}
for ws in wb.worksheets:
    if ws.title in SKIP: continue
    a1 = ws['A1'].value or ws.title
    key = norm(a1)
    sheet_norm_to_title[key] = ws.title
    # also index by tab name itself
    sheet_norm_to_title.setdefault(norm(ws.title), ws.title)

def find_sheet_for_university(uni_name):
    key = norm(uni_name)
    if key in sheet_norm_to_title:
        return sheet_norm_to_title[key]
    # containment fallback
    for k, title in sheet_norm_to_title.items():
        if k and (k in key or key in k):
            return title
    return None

def fmt_val(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if v is None:
        return ''
    return str(v).strip()

# ---- Extract course indexes ----
def extract_course_sheet(name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    header = [fmt_val(h).upper() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r[0]:
            continue
        out.append({
            'course': fmt_val(r[0]),
            'level': fmt_val(r[1]),
            'intake': fmt_val(r[2]),
            'university': fmt_val(r[3]),
            'campus': fmt_val(r[4]) if len(r) > 4 else ''
        })
    return out

ug_courses = extract_course_sheet('UG_COURSES')
pg_courses = extract_course_sheet('PG_COURSES')

# ---- Extract per-university meta + own course list ----
def extract_university_sheet(title):
    ws = wb[title]
    rows = list(ws.iter_rows(values_only=True))
    name = fmt_val(rows[0][0]) if rows and rows[0][0] else title

    # find category header row: a row (within first 6 rows) with multiple non-empty cells after col A,
    # excluding a pure "DOWNLOAD FLYER" or "DEADLINES" row
    cat_row_idx = None
    for i in range(min(6, len(rows))):
        r = rows[i]
        vals = [fmt_val(c) for c in r]
        nonempty = [v for v in vals[1:6] if v]
        if len(nonempty) >= 1 and ('RETURN TO MAIN PAGE' in vals[0].upper() if vals[0] else False):
            cat_row_idx = i
            break
    categories = []
    if cat_row_idx is not None:
        r = rows[cat_row_idx]
        for c in r[1:]:
            v = fmt_val(c)
            if v:
                categories.append(v)
            else:
                break

    meta = {}  # label -> {category: value}
    course_header_idx = None
    i = (cat_row_idx + 1) if cat_row_idx is not None else 3
    while i < len(rows):
        r = rows[i]
        label = fmt_val(r[0]).upper()
        if not label:
            i += 1
            # check if next non-empty row looks like course header
            continue
        if 'COURSE' in label and ('LEVEL' in ''.join(fmt_val(c).upper() for c in r) or True):
            # heuristic: row where col A contains COURSE and col B looks like LEVEL header
            if len(r) > 1 and 'LEVEL' in fmt_val(r[1]).upper():
                course_header_idx = i
                break
        vals = {}
        for j, cat in enumerate(categories, start=1):
            if j < len(r):
                v = fmt_val(r[j])
                if v:
                    vals[cat] = v
        if vals:
            meta[fmt_val(r[0])] = vals
        i += 1

    courses = []
    if course_header_idx is not None:
        header = [fmt_val(c).upper() for c in rows[course_header_idx]]
        for r in rows[course_header_idx+1:]:
            if not fmt_val(r[0]):
                continue
            row_label = fmt_val(r[0])
            level_val = fmt_val(r[1]) if len(r) > 1 else ''
            if not level_val and not (len(r) > 2 and fmt_val(r[2])):
                # sub-header row like "LOWER TIER COURSES", skip but keep as tag
                continue
            entry = {
                'course': row_label,
                'level': level_val,
                'intake': fmt_val(r[2]) if len(r) > 2 else '',
            }
            if len(r) > 3 and fmt_val(r[3]):
                entry['campus_or_extra'] = fmt_val(r[3])
            courses.append(entry)

    return {
        'name': name,
        'sheet': title,
        'categories': categories,
        'meta': meta,
        'courses': courses,
    }

universities = {}
for ws in wb.worksheets:
    if ws.title in SKIP: continue
    data = extract_university_sheet(ws.title)
    universities[data['name']] = data

# attach matched university key (from universities dict) to each course row
uni_norm_index = {}
for key in universities:
    uni_norm_index[norm(key)] = key

def match_uni(name):
    k = norm(name)
    if k in uni_norm_index:
        return uni_norm_index[k]
    for nk, orig in uni_norm_index.items():
        if nk and (nk in k or k in nk):
            return orig
    return None

for c in ug_courses:
    c['uni_key'] = match_uni(c['university'])
for c in pg_courses:
    c['uni_key'] = match_uni(c['university'])

unmatched_ug = sum(1 for c in ug_courses if not c['uni_key'])
unmatched_pg = sum(1 for c in pg_courses if not c['uni_key'])
print('unmatched UG:', unmatched_ug, 'unmatched PG:', unmatched_pg)

json.dump(ug_courses, open('/home/claude/site/data/ug_courses.json','w'), ensure_ascii=False)
json.dump(pg_courses, open('/home/claude/site/data/pg_courses.json','w'), ensure_ascii=False)
json.dump(universities, open('/home/claude/site/data/universities.json','w'), ensure_ascii=False)

print('UG courses:', len(ug_courses))
print('PG courses:', len(pg_courses))
print('Universities parsed:', len(universities))
for k,v in list(universities.items())[:3]:
    print('---', k, 'categories=', v['categories'], 'meta_labels=', list(v['meta'].keys())[:8], 'own_courses=', len(v['courses']))
